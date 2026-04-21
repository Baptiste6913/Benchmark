"""Tests de non-régression pour le générateur xlsx.

Génère le bench LISI (bench.json + grille lisi-souverainete.json) et vérifie :
- 3 onglets présents (Matrice / Détails / Sources)
- nombre d'acteurs cohérent avec le bench
- scores pondérés cohérents avec bench.exec_summary.ranking
- somme des poids par acteur = 1.0
- nombre de sources cohérent
- headers attendus sur chaque onglet
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from lib._common import BenchContext  # noqa: E402
from lib.render import render_xlsx  # noqa: E402

LISI_BENCH = REPO_ROOT / "examples" / "generic-demo" / "bench.json"


@pytest.fixture(scope="module")
def rendered_xlsx(tmp_path_factory) -> Path:
    out = tmp_path_factory.mktemp("render") / "lisi.xlsx"
    path = render_xlsx(LISI_BENCH, out)
    assert path.exists()
    return path


@pytest.fixture(scope="module")
def bench_ctx() -> BenchContext:
    return BenchContext.load(LISI_BENCH)


def test_three_sheets(rendered_xlsx: Path) -> None:
    wb = load_workbook(rendered_xlsx)
    assert wb.sheetnames == ["Matrice", "Détails", "Sources"]


def test_matrice_headers(rendered_xlsx: Path, bench_ctx: BenchContext) -> None:
    wb = load_workbook(rendered_xlsx)
    ws = wb["Matrice"]
    # Row 5 = headers
    headers = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers[0] == "#"
    assert headers[1] == "Acteur"
    # Depuis hotfix item 3 : header utilise label_fr (fallback sur name)
    for i, crit in enumerate(bench_ctx.criteria):
        expected = crit.get("label_fr") or crit.get("name") or crit["id"]
        assert headers[2 + i] == expected
    assert "Score pondéré" in headers[-2]
    assert "Verdict" in headers[-1]


def test_matrice_actor_count(rendered_xlsx: Path, bench_ctx: BenchContext) -> None:
    wb = load_workbook(rendered_xlsx)
    ws = wb["Matrice"]
    # Rows 6..5+N = acteurs
    n_actors = len(bench_ctx.actors_by_rank)
    for i in range(n_actors):
        row = 6 + i
        # Col 1 = rank, col 2 = name
        assert ws.cell(row=row, column=1).value == bench_ctx.actors_by_rank[i]["rank"]
        assert ws.cell(row=row, column=2).value == bench_ctx.actors_by_rank[i]["name"]


def test_matrice_weighted_score_matches_bench(rendered_xlsx: Path, bench_ctx: BenchContext) -> None:
    wb = load_workbook(rendered_xlsx)
    ws = wb["Matrice"]
    n_criteria = len(bench_ctx.criteria)
    for i, actor in enumerate(bench_ctx.actors_by_rank):
        row = 6 + i
        cell_value = ws.cell(row=row, column=3 + n_criteria).value
        assert abs(cell_value - actor["weighted_score"]) < 0.01, (
            f"{actor['id']}: xlsx={cell_value} vs bench={actor['weighted_score']}"
        )


def test_weights_sum_to_one(bench_ctx: BenchContext) -> None:
    for actor in bench_ctx.actors_by_rank:
        weights = [r.get("weight") for r in actor.get("scoring", [])]
        total = sum(w for w in weights if w is not None)
        assert abs(total - 1.0) < 0.01, f"{actor['id']} sum(weights) = {total}"


def test_details_row_count(rendered_xlsx: Path, bench_ctx: BenchContext) -> None:
    wb = load_workbook(rendered_xlsx)
    ws = wb["Détails"]
    expected_rows = 4 + len(bench_ctx.actors_by_rank) * len(bench_ctx.criteria)
    assert ws.max_row == expected_rows


def test_details_headers(rendered_xlsx: Path) -> None:
    wb = load_workbook(rendered_xlsx)
    ws = wb["Détails"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, 6)]
    assert headers == ["Acteur", "Critère", "Poids", "Score /5", "Justification"]


def test_sources_count(rendered_xlsx: Path, bench_ctx: BenchContext) -> None:
    wb = load_workbook(rendered_xlsx)
    ws = wb["Sources"]
    expected_rows = 4 + len(bench_ctx.sources_by_id)
    assert ws.max_row == expected_rows


def test_sources_have_urls(rendered_xlsx: Path) -> None:
    wb = load_workbook(rendered_xlsx)
    ws = wb["Sources"]
    # Quelques sources doivent avoir un hyperlink
    found = 0
    for row in range(5, ws.max_row + 1):
        cell = ws.cell(row=row, column=5)
        if cell.hyperlink is not None:
            found += 1
    assert found > 0


def test_render_xlsx_resolves_grid_from_bench() -> None:
    """Confirme que render_xlsx trouve la grille via bench.grid_ref sans argument grid."""
    bench = json.loads(LISI_BENCH.read_text(encoding="utf-8"))
    assert "grid_ref" in bench
    # Le fixture rendered_xlsx l'a déjà testé indirectement.
