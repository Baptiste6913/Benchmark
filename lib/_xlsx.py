"""Générateur xlsx à partir d'un BenchContext.

3 onglets selon la spec S2.2 :
- Matrice : acteurs en lignes, critères pondérés en colonnes + score + verdict
- Détails : 1 ligne par (acteur × critère) avec la justification
- Sources : 1 ligne par source avec URL cliquable

Les couleurs et polices suivent la charte Ascend (voir _common.ASCEND_PALETTE).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from lib._common import (
    ASCEND_FONTS,
    ASCEND_PALETTE,
    BenchContext,
    criterion_display,
    score_color,
    verdict_palette,
)


# -----------------------------------------------------------------------------
# Styles
# -----------------------------------------------------------------------------

BODY_FONT_NAME = ASCEND_FONTS["body"][0]      # Söhne — fallback Calibri si absent sur le poste
TITLE_FONT_NAME = ASCEND_FONTS["title"][0]    # Signifier — fallback Georgia

HEADER_FILL = PatternFill("solid", fgColor=ASCEND_PALETTE["title_primary"])
HEADER_FONT = Font(name=BODY_FONT_NAME, size=11, bold=True, color=ASCEND_PALETTE["white"])
TITLE_FONT = Font(name=TITLE_FONT_NAME, size=14, bold=True, color=ASCEND_PALETTE["title_primary"])
SUBTITLE_FONT = Font(name=BODY_FONT_NAME, size=10, italic=True, color=ASCEND_PALETTE["brun_clair"])
BODY_FONT = Font(name=BODY_FONT_NAME, size=10, color=ASCEND_PALETTE["title_primary"])
BOLD_FONT = Font(name=BODY_FONT_NAME, size=10, bold=True, color=ASCEND_PALETTE["title_primary"])

THIN_BORDER = Border(
    left=Side(border_style="thin", color=ASCEND_PALETTE["accent_secondary"]),
    right=Side(border_style="thin", color=ASCEND_PALETTE["accent_secondary"]),
    top=Side(border_style="thin", color=ASCEND_PALETTE["accent_secondary"]),
    bottom=Side(border_style="thin", color=ASCEND_PALETTE["accent_secondary"]),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)


# -----------------------------------------------------------------------------
# Onglet 1 — Matrice
# -----------------------------------------------------------------------------


def _build_matrice(ws, ctx: BenchContext) -> None:
    ws.title = "Matrice"
    n_criteria = len(ctx.criteria)
    last_col = 2 + n_criteria + 2          # rang, acteur, critères..., pondéré, verdict

    # Titre + sous-titre
    ws["A1"] = f"Bench — {ctx.mission()}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    ws["A2"] = (
        f"Client : {ctx.client()} | Date : {ctx.date()} | Version bench : {ctx.version()} | "
        f"Mode : {ctx.mode()} | Grille : {ctx.grid_path.name} | Consultant : {ctx.consultant()}"
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)

    # Ligne 4 : pondérations
    ws.cell(row=4, column=2, value="Pondération").font = BOLD_FONT
    ws.cell(row=4, column=2).alignment = LEFT_CENTER
    for i, crit in enumerate(ctx.criteria):
        cell = ws.cell(row=4, column=3 + i, value=f"{int(round(crit['weight'] * 100))}%")
        cell.font = BOLD_FONT
        cell.alignment = CENTER

    # Ligne 5 : headers (utilise label_fr > name > id)
    headers = ["#", "Acteur"] + [criterion_display(c) for c in ctx.criteria] + ["Score pondéré /5", "Verdict red-team"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Lignes acteurs (6 onwards)
    for ridx, actor in enumerate(ctx.actors_by_rank, start=6):
        ws.cell(row=ridx, column=1, value=actor.get("rank", "")).alignment = CENTER
        ws.cell(row=ridx, column=2, value=actor.get("name", actor.get("id", ""))).font = BOLD_FONT
        ws.cell(row=ridx, column=2).alignment = LEFT_CENTER

        for i, cid in enumerate(ctx.criterion_ids()):
            score = ctx.actor_score(actor, cid)
            cell = ws.cell(row=ridx, column=3 + i, value=score)
            cell.alignment = CENTER
            cell.font = BODY_FONT
            if score is not None:
                cell.fill = PatternFill("solid", fgColor=score_color(score))

        ws_score = ws.cell(row=ridx, column=3 + n_criteria, value=round(actor.get("weighted_score", 0.0), 2))
        ws_score.alignment = CENTER
        ws_score.font = BOLD_FONT
        ws_score.fill = PatternFill("solid", fgColor=score_color(actor.get("weighted_score")))

        verdict = actor.get("red_team_verdict") or "—"
        pal = verdict_palette(verdict)
        vcell = ws.cell(row=ridx, column=3 + n_criteria + 1, value=verdict)
        vcell.alignment = CENTER
        vcell.font = Font(name=BODY_FONT_NAME, size=10, bold=True, color=pal["fg"])
        vcell.fill = PatternFill("solid", fgColor=pal["bg"])

        for c in range(1, last_col + 1):
            ws.cell(row=ridx, column=c).border = THIN_BORDER

    # Bloc "Positionnement synthétique" après un saut de ligne
    start_pos = 6 + len(ctx.actors_by_rank) + 2
    ws.cell(row=start_pos, column=1, value="Positionnement synthétique").font = TITLE_FONT
    ws.merge_cells(start_row=start_pos, start_column=1, end_row=start_pos, end_column=last_col)

    for offset, actor in enumerate(ctx.actors_by_rank, start=1):
        r = start_pos + offset
        ws.cell(row=r, column=1, value=f"#{actor.get('rank', '')}").font = BOLD_FONT
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2, value=actor.get("name", "")).font = BOLD_FONT
        ws.cell(row=r, column=2).alignment = LEFT_CENTER
        positioning = (
            actor.get("positioning_one_liner")
            or (actor.get("tldr") or {}).get("positioning", "")
        )
        pc = ws.cell(row=r, column=3, value=positioning)
        pc.alignment = LEFT_TOP
        pc.font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=last_col)
        ws.row_dimensions[r].height = 35

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    for i in range(n_criteria):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.column_dimensions[get_column_letter(3 + n_criteria)].width = 16
    ws.column_dimensions[get_column_letter(3 + n_criteria + 1)].width = 22

    # Freeze panes : acteurs figés à gauche, en-têtes en haut
    ws.freeze_panes = ws.cell(row=6, column=3).coordinate

    # Auto-filter sur la matrice scoring
    first_col = get_column_letter(1)
    last_col_letter = get_column_letter(last_col)
    last_data_row = 5 + len(ctx.actors_by_rank)
    ws.auto_filter.ref = f"{first_col}5:{last_col_letter}{last_data_row}"


# -----------------------------------------------------------------------------
# Onglet 2 — Détails
# -----------------------------------------------------------------------------


def _build_details(wb, ctx: BenchContext) -> None:
    ws = wb.create_sheet("Détails")

    ws["A1"] = "Justifications détaillées — scoring par acteur et critère"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = (
        "Chaque justification s'appuie sur la fiche investigator sourcée "
        "et doit être lue avec le bench.json complet."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:E2")

    headers = ["Acteur", "Critère", "Poids", "Score /5", "Justification"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    row = 5
    for actor in ctx.actors_by_rank:
        for crit in ctx.criteria:
            score = ctx.actor_score(actor, crit["id"])
            justif = ctx.actor_justification(actor, crit["id"])

            ws.cell(row=row, column=1, value=actor.get("name", "")).font = BOLD_FONT
            ws.cell(row=row, column=1).alignment = LEFT_TOP

            ws.cell(row=row, column=2, value=criterion_display(crit)).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = LEFT_TOP

            wc = ws.cell(row=row, column=3, value=f"{int(round(crit['weight'] * 100))}%")
            wc.font = BODY_FONT
            wc.alignment = CENTER

            sc = ws.cell(row=row, column=4, value=score)
            sc.font = BOLD_FONT
            sc.alignment = CENTER
            if score is not None:
                sc.fill = PatternFill("solid", fgColor=score_color(score))

            jc = ws.cell(row=row, column=5, value=justif)
            jc.font = BODY_FONT
            jc.alignment = LEFT_TOP

            for c in range(1, 6):
                ws.cell(row=row, column=c).border = THIN_BORDER

            ws.row_dimensions[row].height = max(30, min(120, 22 + len(justif) // 18))
            row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 95

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{row - 1}"


# -----------------------------------------------------------------------------
# Onglet 3 — Sources
# -----------------------------------------------------------------------------


_RELIABILITY_PALETTE = {
    5: "CFE0C3",  # vert doux
    4: "E4EED5",
    3: "F2DAB8",  # orange doux
    2: "F4C7C1",
    1: "F4C7C1",  # rouge doux
}


def _build_sources(wb, ctx: BenchContext) -> None:
    ws = wb.create_sheet("Sources")

    ws["A1"] = "Sources — panel complet"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        f"Total : {len(ctx.sources_by_id)} sources. "
        "Priorité aux sources primaires et à la triangulation sur les chiffres critiques."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:G2")

    headers = ["Acteur", "Titre", "Média / Auteur", "Date", "URL", "Type", "Fiabilité /5"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Lookup actor name from id for readability
    actor_name_by_id = {a["id"]: a.get("name", a["id"]) for a in ctx.bench.get("actors", [])}

    # Sort sources : by actor (in rank order) then by source id
    rank_by_actor_id = {a["id"]: a.get("rank", 999) for a in ctx.actors_by_rank}

    def source_sort_key(s: dict):
        ar = s.get("actor_ref", "global")
        return (rank_by_actor_id.get(ar, 999), ar, s.get("id", ""))

    sorted_sources = sorted(ctx.bench.get("sources", []), key=source_sort_key)

    for idx, src in enumerate(sorted_sources, start=5):
        actor_ref = src.get("actor_ref", "global")
        actor_label = actor_name_by_id.get(actor_ref, actor_ref)

        ws.cell(row=idx, column=1, value=actor_label).font = BOLD_FONT
        ws.cell(row=idx, column=1).alignment = LEFT_TOP

        ws.cell(row=idx, column=2, value=src.get("title", "")).font = BODY_FONT
        ws.cell(row=idx, column=2).alignment = LEFT_TOP

        ws.cell(row=idx, column=3, value=src.get("author", "")).font = BODY_FONT
        ws.cell(row=idx, column=3).alignment = LEFT_TOP

        dc = ws.cell(row=idx, column=4, value=src.get("date", ""))
        dc.font = BODY_FONT
        dc.alignment = CENTER

        url = src.get("url", "")
        uc = ws.cell(row=idx, column=5, value=url)
        if url:
            uc.hyperlink = url
            uc.font = Font(name=BODY_FONT_NAME, size=9, color="0563C1", underline="single")
        else:
            uc.font = BODY_FONT
        uc.alignment = LEFT_TOP

        tc = ws.cell(row=idx, column=6, value=src.get("type", ""))
        tc.font = BODY_FONT
        tc.alignment = CENTER

        rel = src.get("reliability")
        fc = ws.cell(row=idx, column=7, value=rel)
        fc.font = BOLD_FONT
        fc.alignment = CENTER
        if isinstance(rel, (int, float)) and int(rel) in _RELIABILITY_PALETTE:
            fc.fill = PatternFill("solid", fgColor=_RELIABILITY_PALETTE[int(rel)])

        for c in range(1, 8):
            ws.cell(row=idx, column=c).border = THIN_BORDER

        ws.row_dimensions[idx].height = 28

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 12

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{4 + len(sorted_sources)}"


# -----------------------------------------------------------------------------
# Point d'entrée
# -----------------------------------------------------------------------------


def _build_exclusions(wb, ctx: BenchContext) -> None:
    """Onglet 'Exclusions dealbreaker' (v1.3.1).

    Créé uniquement si exec_summary.exclus_dealbreaker est non vide. Ne
    perturbe pas les benchs pré-v1.3 (pas de feuille supplémentaire).
    """
    excluded_ids = (ctx.bench.get("exec_summary") or {}).get("exclus_dealbreaker") or []
    if not excluded_ids:
        return

    actor_by_id = {a["id"]: a for a in ctx.bench.get("actors", [])}
    excluded_actors = [actor_by_id[aid] for aid in excluded_ids if aid in actor_by_id]
    if not excluded_actors:
        return

    ws = wb.create_sheet("Exclusions")

    ws["A1"] = "Acteurs écartés — dealbreaker Phase 0"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "Écartés avant scoring pondéré. Chaque violation référence une règle "
        "du scope.yaml.dealbreakers avec preuve factuelle."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:D2")

    headers = ["Acteur", "Règle violée", "Preuve", "Source ID"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    row = 5
    for actor in excluded_actors:
        violations = actor.get("dealbreaker_violations", []) or []
        if not violations:
            ws.cell(row=row, column=1, value=actor.get("name", actor["id"])).font = BOLD_FONT
            ws.cell(row=row, column=2, value="(violations non détaillées)").font = BODY_FONT
            for c in range(1, 5):
                cell = ws.cell(row=row, column=c)
                cell.border = THIN_BORDER
                cell.alignment = LEFT_TOP
            row += 1
            continue

        for v in violations:
            ws.cell(row=row, column=1, value=actor.get("name", actor["id"])).font = BOLD_FONT
            ws.cell(row=row, column=2, value=v.get("rule_id", "")).font = BODY_FONT
            ws.cell(row=row, column=3, value=v.get("evidence", "")).font = BODY_FONT
            ws.cell(row=row, column=4, value=v.get("source_id", "")).font = BODY_FONT
            for c in range(1, 5):
                cell = ws.cell(row=row, column=c)
                cell.border = THIN_BORDER
                cell.alignment = LEFT_TOP
                cell.fill = PatternFill("solid", fgColor="F4C7C1")  # rouge doux — exclusion
            ws.row_dimensions[row].height = 35
            row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 16

    ws.freeze_panes = "A5"


def write_xlsx(ctx: BenchContext, out_path: Path) -> None:
    wb = Workbook()
    _build_matrice(wb.active, ctx)
    _build_details(wb, ctx)
    _build_sources(wb, ctx)
    # v1.3.1 : onglet Exclusions conditionnel (absent si pas de dealbreaker violated)
    _build_exclusions(wb, ctx)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
